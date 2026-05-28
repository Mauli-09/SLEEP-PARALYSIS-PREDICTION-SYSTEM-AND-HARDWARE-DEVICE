"""
=============================================================================
  SLEEP PARALYSIS — DATASET C FINAL GENERATOR
  Input  1 : Raw_Dataset_Realistic_60days.xlsx
  Input  2 : REM_Timeline_Realistic_60days.xlsx
  Output   : Dataset_C_Final_60days.xlsx
=============================================================================
"""
import math, os, warnings
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
warnings.filterwarnings('ignore')

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
INPUT_RAW   = '/mnt/user-data/outputs/Raw_Dataset_Realistic_60days.xlsx'
INPUT_REM   = '/mnt/user-data/outputs/REM_Timeline_Realistic_60days.xlsx'
OUTPUT_DIR  = '/mnt/user-data/outputs'
OUTPUT_FILE = 'Dataset_C_Final_60days.xlsx'
SLOT_MIN    = 15
SLOTS_HR    = 4
TARGET_SLEEP= 8.0


# =============================================================================
#  UTILITIES
# =============================================================================
def time_to_dec(t):
    if not t or pd.isna(t) or str(t).strip() == '': return None
    h, m = map(int, str(t).strip().split(':'))
    return h + m / 60.0

def hrs_into_sleep(clock, start):
    return (clock - start) % 24

def in_rem(clock, windows):
    for rs, re in windows:
        if rs <= re:
            if rs <= clock < re: return True
        else:
            if clock >= rs or clock < re: return True
    return False

def cyc_encode(hour_dec):
    s = math.sin(2 * math.pi * hour_dec / 24)
    c = math.cos(2 * math.pi * hour_dec / 24)
    return round(s, 6), round(c, 6)

def parse_date(d):
    if isinstance(d, str): return datetime.strptime(d, '%Y-%m-%d').date()
    if hasattr(d, 'date'): return d.date()
    return d


# =============================================================================
#  LOAD
# =============================================================================
print("=" * 65)
print("  DATASET C GENERATOR — 60-DAY REALISTIC")
print("=" * 65)
print("\n[1/6] Loading files...")

df_raw = pd.read_excel(INPUT_RAW)
df_rem = pd.read_excel(INPUT_REM)
df_raw['Date'] = pd.to_datetime(df_raw['Date']).dt.date
df_rem['Date'] = pd.to_datetime(df_rem['Date']).dt.date

df_raw['_ss']  = df_raw['Sleep_Start'].apply(time_to_dec)
df_raw['_atk'] = df_raw['Attack_Time'].apply(time_to_dec)
df_rem['_rs']  = df_rem['REM_Start'].apply(time_to_dec)
df_rem['_re']  = df_rem['REM_End'].apply(time_to_dec)

PATIENTS = sorted(df_raw['Patient_ID'].unique())
print(f"    Raw  : {df_raw.shape[0]} rows | {df_raw['Attack'].sum()} attacks")
print(f"    REM  : {df_rem.shape[0]} rows")
print(f"    Patients : {PATIENTS}")


# =============================================================================
#  MAIN LOOP
# =============================================================================
print("\n[2/6] Engineering slot-level features...")
rows = []

for pid in PATIENTS:
    print(f"\n    {pid}...")
    pat_raw = df_raw[df_raw['Patient_ID']==pid].sort_values('Date').reset_index(drop=True)
    pat_rem = df_rem[df_rem['Patient_ID']==pid]

    attack_log   = []   # (date, attack_time_dec)
    dur_log      = []   # (date, sleep_duration) for debt

    for _, day in pat_raw.iterrows():
        cur_date   = parse_date(day['Date'])
        ss         = day['_ss']
        dur        = float(day['Sleep_Duration'])
        sleep_eff  = float(day['Sleep_Efficiency'])
        hrv        = float(day['HRV_ms'])
        stress     = int(day['Stress'])
        anxiety    = int(day['Anxiety_Score'])
        caffeine   = float(day['Caffeine_hrs_before_bed'])
        alcohol    = int(day['Alcohol'])
        has_atk    = int(day['Attack']) == 1
        atk_dec    = day['_atk']

        dur_log.append((cur_date, dur))

        # REM windows tonight
        night_rem = pat_rem[pat_rem['Date'] == cur_date]
        rem_wins  = [
            (r['_rs'], r['_re'])
            for _, r in night_rem.iterrows()
            if pd.notna(r['_rs']) and pd.notna(r['_re'])
        ]

        # Attack slot index
        atk_slot = None
        if has_atk and atk_dec is not None:
            atk_slot = int(hrs_into_sleep(atk_dec, ss) * SLOTS_HR)
            atk_slot = min(atk_slot, int(dur * SLOTS_HR) - 1)

        # Past attacks (before tonight — no leakage)
        past = [(d, t) for d, t in attack_log if d < cur_date]
        n_past = len(past)

        # ── 1. days_since_last_attack ──────────────────────────────────────
        days_since = 999 if not past else (cur_date - past[-1][0]).days

        # ── 2. SPHI — attacks / 30 days ───────────────────────────────────
        a30   = [x for x in past if (cur_date - x[0]).days <= 30]
        sphi  = round(len(a30) / 30, 6)

        # ── 3. rolling_attack_mean_7 ───────────────────────────────────────
        a7     = [x for x in past if (cur_date - x[0]).days <= 7]
        roll7  = round(len(a7) / 7, 6)

        # ── 4. Mean_Attack_Interval ────────────────────────────────────────
        if n_past >= 2:
            ivs   = [(past[i][0]-past[i-1][0]).days for i in range(1, n_past)]
            m_int = round(float(np.mean(ivs)), 2)
        elif n_past == 1:
            m_int = float((cur_date - past[0][0]).days)
        else:
            m_int = 0.0

        # ── 5. Attack_Time_STD ─────────────────────────────────────────────
        ptimes   = [t for _, t in past]
        atk_std  = round(float(np.std(ptimes)*60), 2) if len(ptimes)>=2 else 0.0

        # ── 6. SPVI — std of sleep start last 7 nights (minutes) ──────────
        ss_7 = [float(row['_ss']) for _, row in pat_raw.iterrows()
                if 0 < (cur_date - parse_date(row['Date'])).days <= 7]
        spvi = round(float(np.std(ss_7)*60), 2) if len(ss_7)>=2 else 0.0

        # ── 7. Sleep_Debt_3d ───────────────────────────────────────────────
        p3   = [s for _, s in dur_log[-3:]]
        debt = round(TARGET_SLEEP - float(np.mean(p3)), 2)

        # ── 8. Dynamic REM_Vulnerability_Index ────────────────────────────
        static_rv = float(day.get('REM_Vulnerability_Index', 0.75))
        rec14     = past[-14:]
        if rec14:
            hits = 0
            for ad, at in rec14:
                ar = pat_rem[pat_rem['Date']==ad]
                aw = [(r['_rs'], r['_re']) for _, r in ar.iterrows() if pd.notna(r['_rs'])]
                if in_rem(at, aw): hits += 1
            rem_vuln = round(hits / len(rec14), 3)
        else:
            rem_vuln = static_rv

        n_past_nights = len([d for d, _ in dur_log if d < cur_date])
        n_slots = int(dur * SLOTS_HR)

        # ── PER SLOT ──────────────────────────────────────────────────────
        for si in range(n_slots):
            clock   = (ss + si * SLOT_MIN / 60.0) % 24
            mins_on = si * SLOT_MIN
            h_sin, h_cos = cyc_encode(clock)
            rem_slot = int(in_rem(clock, rem_wins))

            # SPTI
            hits_slot = 0
            for ad, at in past:
                arow = pat_raw[pat_raw['Date']==ad]
                if len(arow)==0: continue
                a_ss  = arow.iloc[0]['_ss']
                a_sl  = int(hrs_into_sleep(at, a_ss) * SLOTS_HR)
                if a_sl == si: hits_slot += 1
            spti = round(hits_slot / max(1, n_past_nights), 6)

            label = 1 if (has_atk and atk_slot == si) else 0

            rows.append({
                'Patient_ID'               : pid,
                'Date'                     : str(cur_date),
                'Slot_Index'               : si,
                'hour_sin'                 : h_sin,
                'hour_cos'                 : h_cos,
                'minutes_since_sleep_onset': mins_on,
                'REM_slot'                 : rem_slot,
                'REM_Vulnerability_Index'  : rem_vuln,
                'days_since_last_attack'   : days_since,
                'SPHI'                     : sphi,
                'rolling_attack_mean_7'    : roll7,
                'SPTI'                     : spti,
                'Mean_Attack_Interval'     : m_int,
                'Attack_Time_STD'          : atk_std,
                'Sleep_Duration'           : dur,
                'Sleep_Efficiency'         : sleep_eff,
                'Sleep_Debt_3d'            : debt,
                'SPVI'                     : spvi,
                'HRV_ms'                   : hrv,
                'Stress'                   : stress,
                'Anxiety_Score'            : anxiety,
                'Caffeine_hrs_before_bed'  : caffeine,
                'Alcohol'                  : alcohol,
                'Label'                    : label,
            })

        if has_atk and atk_dec is not None:
            attack_log.append((cur_date, atk_dec))

    pat_atks = sum(1 for d,_ in attack_log)
    pat_rows = sum(1 for r in rows if r['Patient_ID']==pid)
    print(f"      {pat_atks} attacks | {pat_rows} slot rows")


# =============================================================================
#  DATAFRAME + VALIDATION
# =============================================================================
print("\n[3/6] Validation...")
df_c  = pd.DataFrame(rows)
total = len(df_c)
pos   = df_c['Label'].sum()
neg   = total - pos
ratio = neg / pos
rem_pct = df_c[df_c['Label']==1]['REM_slot'].mean()*100

print(f"\n    ✅ Validation Report")
print(f"       Rows                : {total:,}")
print(f"       Label=1 (attack)    : {pos}")
print(f"       Label=0 (normal)    : {neg:,}")
print(f"       Imbalance ratio     : {ratio:.1f}:1")
print(f"       scale_pos_weight    : {ratio:.1f}")
print(f"       REM attack slots    : {rem_pct:.1f}%  (target >85%)")
print(f"       days_since=999 ok   : {(df_c.groupby('Patient_ID').first()['days_since_last_attack']==999).all()}")
print(f"\n    Attacks per patient:")
for p in PATIENTS:
    print(f"       {p}: {df_c[df_c['Patient_ID']==p]['Label'].sum()}")

# Correlation check on slot-level data
print(f"\n    Key correlations (slot level):")
print(f"    {'Feature':<30} {'Corr with Label':>16}")
print(f"    {'-'*48}")
for feat in ['Anxiety_Score','HRV_ms','Sleep_Debt_3d','SPTI',
             'days_since_last_attack','minutes_since_sleep_onset','REM_slot']:
    r = df_c[feat].corr(df_c['Label'])
    print(f"    {feat:<30} {r:>+16.4f}")


# =============================================================================
#  TRAIN / TEST SPLIT (chronological 80/20)
# =============================================================================
print("\n[4/6] Train/test split (chronological)...")
dates_sorted = sorted(df_c['Date'].unique())
split_date   = dates_sorted[int(len(dates_sorted)*0.8)]
df_train = df_c[df_c['Date'] <  split_date].copy()
df_test  = df_c[df_c['Date'] >= split_date].copy()
df_c['Split'] = df_c['Date'].apply(lambda d: 'train' if d < split_date else 'test')
print(f"    Split date  : {split_date}")
print(f"    Train       : {len(df_train):,} rows | {df_train['Label'].sum()} attacks")
print(f"    Test        : {len(df_test):,} rows  | {df_test['Label'].sum()} attacks")


# =============================================================================
#  EXPORT
# =============================================================================
print("\n[5/6] Exporting...")

xgb_starter = pd.DataFrame([
    ['# ── STEP 1: Drop non-feature columns',
     "DROP = ['Patient_ID','Date','Slot_Index','Split','Label']"],
    ['# ── STEP 2: Define X and y',
     "FEATURES = [c for c in df_train.columns if c not in DROP]"],
    ['', "X_train, y_train = df_train[FEATURES], df_train['Label']"],
    ['', "X_test,  y_test  = df_test[FEATURES],  df_test['Label']"],
    [f'# ── STEP 3: Class imbalance  scale_pos_weight={ratio:.1f}',
     'from xgboost import XGBClassifier'],
    ['', 'model = XGBClassifier('],
    ['', '    n_estimators      = 300,'],
    ['', '    max_depth         = 5,'],
    ['', '    learning_rate     = 0.05,'],
    ['', f'    scale_pos_weight  = {ratio:.1f},'],
    ['', '    subsample         = 0.8,'],
    ['', '    colsample_bytree  = 0.8,'],
    ['', "    eval_metric       = 'auc',"],
    ['', '    random_state      = 42'],
    ['', ')'],
    ['', 'model.fit(X_train, y_train)'],
    ['# ── STEP 4: Evaluate with AUC + F1 (not accuracy)',
     'from sklearn.metrics import roc_auc_score, f1_score, classification_report'],
    ['', 'y_prob = model.predict_proba(X_test)[:,1]'],
    ['', 'y_pred = (y_prob > 0.55).astype(int)'],
    ['', "print('AUC :', roc_auc_score(y_test, y_prob))"],
    ['', "print('F1  :', f1_score(y_test, y_pred))"],
    ['', 'print(classification_report(y_test, y_pred))'],
    ['# ── STEP 5: Feature importance',
     'import matplotlib.pyplot as plt'],
    ['', 'from xgboost import plot_importance'],
    ['', 'plot_importance(model, max_num_features=15)'],
    ['', 'plt.tight_layout(); plt.show()'],
    ['# ── STEP 6: Predict risk slots for next night',
     "THRESHOLD = 0.55"],
    ['', "df_test['Risk_Prob'] = y_prob"],
    ['', "high_risk = df_test[df_test['Risk_Prob']>THRESHOLD][['Date','Slot_Index','Risk_Prob']]"],
    ['', 'print(high_risk)'],
], columns=['Comment', 'Code'])

path_out = os.path.join(OUTPUT_DIR, OUTPUT_FILE)
with pd.ExcelWriter(path_out, engine='openpyxl') as w:
    df_c.to_excel(w,        index=False, sheet_name='Dataset_C_Slots')
    df_train.to_excel(w,    index=False, sheet_name='Train_Set')
    df_test.to_excel(w,     index=False, sheet_name='Test_Set')
    xgb_starter.to_excel(w, index=False, sheet_name='XGBoost_Starter_Code')


# ── Formatting ────────────────────────────────────────────────────────────────
wb   = load_workbook(path_out)
thin = Side(style='thin', color='CCCCCC')
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

def fmt_sheet(ws, cols_list, label_col='Label', rem_col='REM_slot'):
    li = cols_list.index(label_col)+1 if label_col in cols_list else None
    ri = cols_list.index(rem_col)+1   if rem_col   in cols_list else None
    for cell in ws[1]:
        cell.fill = PatternFill('solid', fgColor='1E3A5F')
        cell.font = Font(bold=True, color='FFFFFF', size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = brd
    ws.row_dimensions[1].height = 32
    for ri2, row in enumerate(ws.iter_rows(min_row=2), 2):
        is_atk = li and row[li-1].value == 1
        is_rem = ri and row[ri-1].value == 1
        base   = ('FF6B6B' if is_atk else ('D5F5E3' if is_rem else
                  ('F0F4FF' if ri2%2==0 else 'FFFFFF')))
        for cell in row:
            cell.fill = PatternFill('solid', fgColor=base)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = brd
            cell.font   = Font(size=8)
    for col in ws.columns:
        mx = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(mx+2,9),22)
    ws.freeze_panes = 'D2'

cols_c = list(df_c.columns)
fmt_sheet(wb['Dataset_C_Slots'], cols_c)
fmt_sheet(wb['Train_Set'],       cols_c)
fmt_sheet(wb['Test_Set'],        cols_c)

# XGBoost sheet style
ws_x = wb['XGBoost_Starter_Code']
for cell in ws_x[1]:
    cell.fill = PatternFill('solid', fgColor='1E3A5F')
    cell.font = Font(bold=True, color='FFFFFF', size=10)
    cell.alignment = Alignment(horizontal='center')
    cell.border = brd
for ri2, row in enumerate(ws_x.iter_rows(min_row=2), 2):
    is_cmt = str(row[0].value or '').startswith('#')
    for ci, cell in enumerate(row):
        cell.fill = PatternFill('solid', fgColor='FFF9C4' if (ci==0 and is_cmt) else 'F8F8F8')
        cell.font = Font(size=9,
                         bold=(ci==0 and is_cmt),
                         color=('E65100' if (ci==0 and is_cmt) else '000000'),
                         name=('Calibri' if ci==0 else 'Courier New'))
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = brd
ws_x.column_dimensions['A'].width = 45
ws_x.column_dimensions['B'].width = 62

wb.save(path_out)

print(f"\n[6/6] ✅  Saved → {path_out}")
print(f"\n{'='*65}")
print(f"  Sheets:")
print(f"  1. Dataset_C_Slots    {len(df_c):>6,} rows  | all patients all days")
print(f"  2. Train_Set          {len(df_train):>6,} rows  | days 1-48")
print(f"  3. Test_Set           {len(df_test):>6,} rows  | days 49-60")
print(f"  4. XGBoost_Starter_Code        copy-paste ready")
print(f"\n  scale_pos_weight = {ratio:.1f}  ← set this in XGBoost")
print(f"  Features ({len(cols_c)-4}): {[c for c in cols_c if c not in ['Patient_ID','Date','Slot_Index','Split','Label']]}")
print(f"{'='*65}")
